from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def _format_currency_columns(ws, headers: list[str]) -> None:
    currency_cols = {"amount", "debit", "credit", "running_balance"}
    for idx, header in enumerate(headers, start=1):
        if header in currency_cols:
            col_letter = get_column_letter(idx)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = "$#,##0.00"


def build_workpaper(
    transactions: pd.DataFrame,
    flags: pd.DataFrame,
    raw_text: pd.DataFrame,
    parsed_at: datetime,
) -> bytes:
    output = BytesIO()
    risk_counts = (
        transactions["risk_level"].value_counts().to_dict()
        if not transactions.empty and "risk_level" in transactions.columns
        else {}
    )
    avg_risk = (
        float(transactions["risk_score"].fillna(0).mean())
        if not transactions.empty and "risk_score" in transactions.columns
        else 0.0
    )

    summary_rows = [
        {"metric": "Parsed Timestamp", "value": parsed_at.strftime("%Y-%m-%d %H:%M:%S")},
        {"metric": "Total Statements", "value": transactions["statement_id"].nunique() if not transactions.empty else 0},
        {"metric": "Total Transactions", "value": len(transactions)},
        {"metric": "Total Debits", "value": float(transactions["debit"].fillna(0).sum()) if not transactions.empty else 0},
        {"metric": "Total Credits", "value": float(transactions["credit"].fillna(0).sum()) if not transactions.empty else 0},
        {"metric": "Average Risk Score", "value": round(avg_risk, 2)},
        {"metric": "High Risk Count", "value": int(risk_counts.get("high", 0))},
        {"metric": "Medium Risk Count", "value": int(risk_counts.get("medium", 0))},
        {"metric": "Low Risk Count", "value": int(risk_counts.get("low", 0))},
        {"metric": "Flag Count", "value": len(flags)},
    ]
    summary = pd.DataFrame(summary_rows)
    risk_summary = pd.DataFrame()
    if not transactions.empty and {"txn_category", "risk_score", "amount"}.issubset(transactions.columns):
        risk_summary = (
            transactions.groupby(["txn_category", "risk_level"], dropna=False)
            .agg(
                txn_count=("txn_category", "size"),
                total_amount=("amount", "sum"),
                avg_risk_score=("risk_score", "mean"),
            )
            .reset_index()
            .sort_values(["avg_risk_score", "txn_count"], ascending=[False, False])
        )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)

        transactions_out = transactions.copy()
        if not transactions_out.empty:
            transactions_out["txn_date"] = transactions_out["txn_date"].dt.strftime("%Y-%m-%d")
        transactions_out.to_excel(writer, sheet_name="Transactions", index=False)

        if not risk_summary.empty:
            risk_summary.to_excel(writer, sheet_name="Risk_Summary", index=False)

        flags.to_excel(writer, sheet_name="Flags_For_Review", index=False)

        if not raw_text.empty:
            raw_text.to_excel(writer, sheet_name="Raw_Text", index=False)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            _style_sheet(ws)
            headers = [cell.value for cell in ws[1]]
            _format_currency_columns(ws, headers)

    return output.getvalue()
