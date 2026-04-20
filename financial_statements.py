from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)


def _format_currency_columns(ws, headers: list[str]) -> None:
    currency_like = {
        "amount",
        "debit",
        "credit",
        "net_cash_change",
        "beginning_balance",
        "ending_balance",
        "expected_ending_balance",
        "reconciliation_difference",
        "value",
    }
    for idx, header in enumerate(headers, start=1):
        if header and str(header).lower() in currency_like:
            col_letter = get_column_letter(idx)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = "$#,##0.00"


def build_financial_statements(
    transactions: pd.DataFrame,
    reconciliation: pd.DataFrame,
    parsed_at: datetime,
) -> bytes:
    output = BytesIO()
    tx = transactions.copy()
    tx["credit"] = pd.to_numeric(tx.get("credit"), errors="coerce").fillna(0.0)
    tx["debit"] = pd.to_numeric(tx.get("debit"), errors="coerce").fillna(0.0)
    tx["signed_amount"] = tx["credit"] - tx["debit"]
    tx["txn_date"] = pd.to_datetime(tx.get("txn_date"), errors="coerce")
    tx["month"] = tx["txn_date"].dt.to_period("M").astype(str)
    tx["txn_category"] = tx.get("txn_category", "other_debit")

    total_inflows = float(tx["credit"].sum())
    total_outflows = float(tx["debit"].sum())
    net_change = total_inflows - total_outflows

    management_summary = pd.DataFrame(
        [
            {"metric": "Prepared Timestamp", "value": parsed_at.strftime("%Y-%m-%d %H:%M:%S")},
            {"metric": "Total Statements", "value": tx["statement_id"].nunique() if not tx.empty else 0},
            {"metric": "Total Transactions", "value": len(tx)},
            {"metric": "Total Cash Inflows", "value": total_inflows},
            {"metric": "Total Cash Outflows", "value": total_outflows},
            {"metric": "Net Cash Change", "value": net_change},
        ]
    )

    activities = pd.DataFrame(columns=["month", "txn_category", "amount"])
    activities_pivot = pd.DataFrame()
    if not tx.empty:
        activities = (
            tx.groupby(["month", "txn_category"], dropna=False)["signed_amount"]
            .sum()
            .reset_index(name="amount")
            .sort_values(["month", "txn_category"])
        )
        activities_pivot = activities.pivot_table(
            index="month",
            columns="txn_category",
            values="amount",
            aggfunc="sum",
            fill_value=0.0,
        ).reset_index()
        value_cols = [c for c in activities_pivot.columns if c != "month"]
        activities_pivot["net_cash_change"] = activities_pivot[value_cols].sum(axis=1)

    cash_position = pd.DataFrame()
    if not reconciliation.empty:
        keep_cols = [
            "statement_id",
            "beginning_balance",
            "total_credits",
            "total_debits",
            "expected_ending_balance",
            "ending_balance",
            "reconciliation_difference",
            "status",
        ]
        available = [c for c in keep_cols if c in reconciliation.columns]
        cash_position = reconciliation[available].copy()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        management_summary.to_excel(writer, sheet_name="Management_Summary", index=False)
        if not activities_pivot.empty:
            activities_pivot.to_excel(writer, sheet_name="Statement_of_Activities", index=False)
            activities.to_excel(writer, sheet_name="Activity_Detail", index=False)
        else:
            pd.DataFrame([{"note": "No transactions available"}]).to_excel(
                writer, sheet_name="Statement_of_Activities", index=False
            )
        if not cash_position.empty:
            cash_position.to_excel(writer, sheet_name="Cash_Position", index=False)
        else:
            pd.DataFrame([{"note": "No reconciled balances available"}]).to_excel(
                writer, sheet_name="Cash_Position", index=False
            )

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            _style_sheet(ws)
            headers = [cell.value for cell in ws[1]]
            _format_currency_columns(ws, headers)

    return output.getvalue()
